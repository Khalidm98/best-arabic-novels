# Google Sheet integration requirements
# Create Google Cloud Console project
# Enable Google Drive API and Google Sheet API
# Add service account and copy json key to project
# https://docs.google.com/spreadsheets/d/1xdzsn2JpKTsjQMJLbskw8KKzhustJw2GAEU2Fs19CO0/edit?usp=sharing

# PDF with arabic characters requirements
# mkdir covers
# copy "Vazirmatn-Regular.ttf" to "site-packages/reportlab/fonts"

import os.path
from pathlib import Path

from bidi.algorithm import get_display
from bs4 import BeautifulSoup
from flask import Flask, abort, jsonify, render_template, request, send_from_directory
from flask.views import MethodView
from openpyxl import Workbook
from openpyxl.styles import Font
import pandas as pd
import pygsheets
from reportlab.graphics import renderPDF
from reportlab.graphics.barcode.qr import QrCodeWidget
from reportlab.graphics.shapes import Drawing
from reportlab.lib.pagesizes import LETTER
from reportlab.pdfbase.pdfmetrics import registerFont, stringWidth
from reportlab.pdfbase.ttfonts import TTFont
from reportlab.pdfgen.canvas import Canvas
import requests
from rtl.reshaper import reshape


def scrap_to_local_sheet():
    base_url = 'https://ar.wikipedia.org'
    url = 'https://ar.wikipedia.org/wiki/قائمة_أفضل_مئة_رواية_عربية'
    response = requests.get(url)
    if response.status_code != 200:
        raise Exception("Something went wrong!")
    pretty = BeautifulSoup(response.text, "lxml")
    table = pretty.find("table", class_="wikitable sortable")

    wb = Workbook()
    sheet = wb.active
    bold = Font(bold=True)
    for col, label in enumerate(table.find_all("th"), 1):
        sheet.cell(1, col, label.text.strip()).font = bold

    for row, row_tag in enumerate((table.find_all("tr"))[1:], 2):
        for col, cell_tag in enumerate(row_tag.find_all("td"), 1):
            cell = sheet.cell(row, col, cell_tag.text.strip())
            a = cell_tag.find("a")
            if a:
                cell.style = "Hyperlink"
                cell.hyperlink = f"{base_url}{a.attrs['href']}"
    wb.save("best_arabic_novels.xlsx")


def scrap_to_google_sheet():
    url = 'https://ar.wikipedia.org/wiki/قائمة_أفضل_مئة_رواية_عربية'
    response = requests.get(url)
    if response.status_code != 200:
        raise Exception("Something went wrong!")
    pretty = BeautifulSoup(response.text, "lxml")
    table = pretty.find("table", class_="wikitable sortable")

    df = pd.DataFrame(columns=[label.text.strip() for label in table.find_all("th")])
    for row_tag in table.find_all("tr")[1:]:
        df.loc[len(df)] = [cell.text.strip() for cell in row_tag.find_all("td")]

    google_client = pygsheets.authorize(service_file="sheets-service-account.json")
    google_sheet = google_client.open("best_arabic_novels")[0]
    google_sheet.set_dataframe(df, (1, 1))
    google_sheet.add_conditional_formatting(
        "A1", "D1", "NOT_BLANK", {"textFormat": {"bold": True}}
    )


def export_novels_covers():
    base_url = 'https://ar.wikipedia.org'
    url = 'https://ar.wikipedia.org/wiki/قائمة_أفضل_مئة_رواية_عربية'
    response = requests.get(url)
    if response.status_code != 200:
        raise Exception("Something went wrong!")
    pretty = BeautifulSoup(response.text, "lxml")
    table = pretty.find("table", class_="wikitable sortable")

    registerFont(TTFont("Vazirmatn", "Vazirmatn-Regular.ttf"))
    for row, row_tag in enumerate((table.find_all("tr"))[1:], 2):
        cells = row_tag.find_all("td")
        rank = cells[0].text.strip()
        novel = cells[1].text.strip()
        author = cells[2].text.strip()
        novel_url = f'{base_url}{cells[1].a.attrs["href"]}'

        pdf = Canvas(
            f"covers/{rank.zfill(3)}. {novel}.pdf",
            initialFontName="Vazirmatn",
            initialFontSize=32,
            pagesize=LETTER,
        )

        qr = QrCodeWidget(novel_url, barHeight=324, barWidth=324, barBorder=0)
        d = Drawing()
        d.add(qr)
        renderPDF.draw(d, pdf, 144, 288)
        pdf.linkURL(novel_url, (144, 288, 468, 612))

        novel_ar = get_display(reshape(novel))
        author_ar = get_display(reshape(author))
        pdf.drawString((LETTER[0] - stringWidth(novel_ar, "Vazirmatn", 32)) / 2, 216, novel_ar)
        pdf.drawString((LETTER[0] - stringWidth(author_ar, "Vazirmatn", 32)) / 2, 144, author_ar)
        pdf.save()


app = Flask(__name__)


@app.route('/')
def home():
    return render_template("home.html")


@app.route('/scrap-to-local-file/')
def local_scrap():
    scrap_to_local_sheet()
    return render_template("success.html")


@app.route('/scrap-to-google-sheets/')
def google_scrap():
    scrap_to_google_sheet()
    return render_template("success.html")


@app.route('/export-novels-covers/')
def covers_export():
    export_novels_covers()
    return render_template("success.html")


class FileAPI(MethodView):
    base_dir = Path(__file__).resolve().parent
    file_name = "best_arabic_novels.xlsx"

    def get(self):
        return send_from_directory(self.base_dir, self.file_name)

    def post(self):
        if not request.files:
            response = jsonify({'details': 'Request is missing files!'})
            response.status_code = 400
            return response

        file = request.files.get('file')
        if not file:
            response = jsonify({'file': 'This field is required!'})
            response.status_code = 400
            return response

        if file.mimetype != "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet":
            response = jsonify({'details': 'Only "xlsx" files are allowed!'})
            response.status_code = 400
            return response

        file.save(os.path.join(self.base_dir, self.file_name))
        response = jsonify({'details': 'Created successfully.'})
        response.status_code = 201
        return response

    def put(self):
        file = request.files.get('file')
        if not file:
            response = jsonify({'details': 'Updated successfully.'})
            return response

        if file.mimetype != "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet":
            response = jsonify({'details': 'Only "xlsx" files are allowed!'})
            response.status_code = 400
            return response

        file.save(os.path.join(self.base_dir, self.file_name))
        response = jsonify({'details': 'Updated successfully.'})
        response.status_code = 201
        return response

    def delete(self):
        path = os.path.join(self.base_dir, self.file_name)
        if not os.path.exists(path):
            abort(404)
        os.remove(path)
        response = jsonify({"details": "Deleted successfully."})
        response.status_code = 204
        return response


app.add_url_rule(
    "/file/",
    view_func=FileAPI.as_view('file_api'),
    methods=["GET", "POST", "PUT", "DELETE"],
)

if __name__ == '__main__':
    app.run()
